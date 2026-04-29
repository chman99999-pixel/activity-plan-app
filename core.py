"""
core.py — 월별 활동계획서 자동 작성 비즈니스 로직

달력형 계획서(.xls)를 파싱하여 이용자별 엑셀 활동계획서(.xlsx)에
활동 내용을 자동으로 입력한다.
"""
import io
import re
from datetime import date

import holidays as kr_holidays
import xlrd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Side, Font
from copy import copy


# ─── 1. 달력 파싱 ───

def _cell_value(sheet, row, col, is_xlsx=False):
    """xlrd(0-indexed) / openpyxl(1-indexed) 공통 셀 값 접근 헬퍼."""
    if is_xlsx:
        val = sheet.cell(row=row + 1, column=col + 1).value
    else:
        val = sheet.cell(row, col).value
    return val


def _sheet_dims(sheet, is_xlsx=False):
    """시트의 (행 수, 열 수)를 반환한다."""
    if is_xlsx:
        return sheet.max_row or 0, sheet.max_column or 0
    return sheet.nrows, sheet.ncols


def _sheet_name(sheet, is_xlsx=False):
    """시트 이름을 반환한다."""
    if is_xlsx:
        return sheet.title
    return sheet.name


def parse_calendar(xls_bytes: bytes):
    """달력 .xls/.xlsx에서 날짜별 활동 내용을 추출한다.

    Returns:
        activities: {날짜(int): [시간대별 문자열, ...]}
        holidays: set of 날짜(int) — 대체공휴일 등
        month: int — 해당 월
        year: int — 해당 연도
    """
    # 파일 형식 감지: ZIP(PK) 헤더면 xlsx, 아니면 xls
    is_xlsx = xls_bytes[:2] == b'PK'

    if is_xlsx:
        wb = load_workbook(io.BytesIO(xls_bytes), read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
    else:
        wb = xlrd.open_workbook(file_contents=xls_bytes)
        sheet = wb.sheet_by_index(0)

    nrows, ncols = _sheet_dims(sheet, is_xlsx)

    # 셀 내용에서 연도/월 추출 (예: "2026년 3월 주간활동 계획서")
    year = None
    month = None
    for row_idx in range(min(5, nrows)):
        for col in range(ncols):
            val = str(_cell_value(sheet, row_idx, col, is_xlsx)).strip()
            if not val:
                continue
            m = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', val)
            if m:
                year = int(m.group(1))
                month = int(m.group(2))
                break
        if year and month:
            break

    if not year:
        year = date.today().year
    if not month:
        # fallback: 시트명에서 숫자 추출
        digits = re.findall(r'\d+', _sheet_name(sheet, is_xlsx))
        for d in digits:
            n = int(d)
            if 1 <= n <= 12:
                month = n
                break

    activities = {}
    holidays = set()

    # 한국 법정 공휴일 자동 추가 (어린이날, 광복절 등)
    kr = kr_holidays.KR(years=year)
    for h_date in kr:
        if h_date.month == month:
            holidays.add(h_date.day)

    row = 0
    while row < nrows:
        row_vals = [str(_cell_value(sheet, row, c, is_xlsx)).strip() for c in range(ncols)]
        is_week_header = any('주차' in v for v in row_vals)

        if is_week_header:
            date_row = None
            time_start_row = None
            for dr in range(row + 1, min(row + 4, nrows)):
                vals = [str(_cell_value(sheet, dr, c, is_xlsx)).strip() for c in range(ncols)]
                if any('시간' in v for v in vals):
                    continue
                has_date = False
                for c in range(ncols):
                    v = str(_cell_value(sheet, dr, c, is_xlsx)).strip()
                    nums = ''.join(ch for ch in v.split('일')[0].split('(')[0] if ch.isdigit())
                    if nums and 1 <= int(nums) <= 31:
                        has_date = True
                        break
                if has_date:
                    date_row = dr
                    time_start_row = dr + 1
                    break

            if date_row is None:
                row += 1
                continue

            # A열에서 시간대 슬롯 동적 감지 (HH:MM~HH:MM 패턴)
            time_slots = []
            for r in range(time_start_row, nrows):
                a_val = str(_cell_value(sheet, r, 0, is_xlsx)).strip()
                if re.match(r'\d{2}:\d{2}~\d{2}:\d{2}', a_val):
                    time_slots.append(a_val)
                else:
                    break

            dates_in_week = {}
            for c in range(ncols):
                v = str(_cell_value(sheet, date_row, c, is_xlsx)).strip()
                if not v:
                    continue
                nums = ''.join(ch for ch in v.split('일')[0].split('(')[0] if ch.isdigit())
                if nums and 1 <= int(nums) <= 31:
                    d = int(nums)
                    dates_in_week[c] = d
                    if '대체공휴일' in v or '공휴일' in v:
                        holidays.add(d)

            for c, d in dates_in_week.items():
                if d in holidays:
                    continue
                first_val = str(_cell_value(sheet, time_start_row, c, is_xlsx)).strip() if time_start_row < nrows else ''
                if '대체공휴일' in first_val or '공휴일' in first_val:
                    holidays.add(d)
                    continue

                day_activities = []
                for slot_idx, slot_time in enumerate(time_slots):
                    r = time_start_row + slot_idx
                    if r >= nrows:
                        break
                    val = str(_cell_value(sheet, r, c, is_xlsx)).strip()
                    if val and val != 'None':
                        day_activities.append(f"{slot_time} {val}")
                    else:
                        if slot_time == '12:00~13:00':
                            day_activities.append(f"{slot_time} 점심식사 및 위생지원")
                        else:
                            day_activities.append(f"{slot_time} ")

                if day_activities:
                    activities[d] = day_activities

            row = time_start_row + len(time_slots)
        else:
            row += 1

    if is_xlsx:
        wb.close()

    return activities, holidays, month, year


# ─── 2. 이용자 감지 ───

def detect_users(xlsx_bytes: bytes) -> list[str]:
    """템플릿 xlsx의 시트명에서 이용자 이름을 추출한다."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    users = []
    for sn in wb.sheetnames:
        # 시트명 끝에서 한글 이름(2~3자) 추출
        # 예: "26.03월 계획서-유정빈", "활동계획서 유정빈", "유정빈" 등
        m = re.search(r'([가-힣]{2,3})\s*$', sn.strip())
        if m:
            users.append(m.group(1))
    wb.close()
    return users


# ─── 3. 행 수 검증 ───

def count_available_rows(xlsx_bytes: bytes) -> int:
    """첫 번째 이용자 시트에서 데이터 입력 가능한 행 수를 확인한다.
    행 9부터 시작, 행 29까지 (30행은 수식)."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    if not wb.sheetnames:
        wb.close()
        return 0
    ws = wb[wb.sheetnames[0]]
    # 기본적으로 9~29행 = 21행이 입력 가능
    # 실제 행 수를 세서 반환
    count = 0
    for row in range(9, 30):  # 행 9~29
        try:
            cell = ws.cell(row=row, column=1)
            count += 1
        except Exception:
            break
    wb.close()
    return count


# ─── 4. 원본 폰트 정보 확인 ───

def get_font_info(xlsx_bytes: bytes):
    """원본 xlsx에서 D열 폰트 정보를 확인한다."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb[wb.sheetnames[0]]
    cell = ws.cell(row=9, column=4)
    font_name = cell.font.name or '맑은 고딕'
    font_size = cell.font.size or 14
    wb.close()
    return font_name, font_size


# ─── 5. 전달 마지막 평일 계산 ───

def _last_weekday_prev_month(year: int, month: int) -> date:
    """해당 월의 전달 마지막 평일을 반환한다.
    예: 2026년 3월 → 2026년 2월의 마지막 평일
    """
    # 전달 계산
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    # 전달 마지막 날부터 거슬러 올라가며 평일 찾기
    import calendar
    last_day = calendar.monthrange(prev_year, prev_month)[1]
    d = date(prev_year, prev_month, last_day)
    while d.weekday() >= 5:  # 토(5), 일(6)
        d = d.replace(day=d.day - 1)
    return d


# ─── 6. 오후 송영 시간 계산 ───

def _get_last_end_hour(day_activities: list[str]) -> int:
    """하루 활동 목록에서 마지막 활동의 종료 시간(시)을 반환한다.
    예: '15:00~16:00 활동' → 16, '16:00~17:00 활동' → 17
    기본값: 16
    """
    last_hour = 16
    for act in day_activities:
        m = re.match(r'\d{2}:\d{2}~(\d{2}):\d{2}', act)
        if m:
            last_hour = int(m.group(1))
    return last_hour


# ─── 6. 활동계획 입력 ───

def fill_sheets(template_bytes: bytes, activities: dict, holidays: set,
                user_config: dict, provider: str, month: int, year: int):
    """이용자별 시트에 활동계획을 입력한다.

    Returns:
        output_bytes: 완성된 xlsx 바이트
        results: 이용자별 결과 목록
        working_days: 활동일 목록
        formulas_ok: 수식 보존 여부
    """
    wb = load_workbook(io.BytesIO(template_bytes))

    font_name, font_size = get_font_info(template_bytes)
    normal_font = InlineFont(rFont=font_name, sz=font_size)
    red_font = InlineFont(rFont=font_name, sz=font_size, color='FFFF0000')

    def make_cell_value(text):
        if '(협)' not in text:
            return text
        parts = []
        segments = text.split('(협)')
        for i, seg in enumerate(segments):
            if seg:
                parts.append(TextBlock(normal_font, seg))
            if i < len(segments) - 1:
                parts.append(TextBlock(red_font, '(협)'))
        return CellRichText(*parts)

    # 활동일 목록 (공휴일 제외, 정렬)
    working_days = sorted([d for d in activities.keys() if d not in holidays])

    # 요일 매핑
    dow_names_kr = ['월', '화', '수', '목', '금', '토', '일']
    dow_map = {}
    for d in working_days:
        try:
            dt = date(year, month, d)
            dow_map[d] = dow_names_kr[dt.weekday()]
        except ValueError:
            dow_map[d] = ''

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 이용자 이름 매칭
        user_name = None
        for name in user_config:
            if name in sheet_name:
                user_name = name
                break
        if not user_name:
            continue

        config = user_config[user_name]
        has_오전송영 = config.get('오전송영', False)
        오전송영시간 = config.get('오전송영시간', '08:30~09:00 송영')
        has_오후송영 = config.get('오후송영', False)
        오후송영시간 = config.get('오후송영시간', '16:00~16:30 송영')
        수급시간 = config.get('수급시간', 132)
        shuttle_count = int(has_오전송영) + int(has_오후송영)
        num_slots = len(next(iter(activities.values()), []))
        row_height = (num_slots + shuttle_count) * 17 + 7

        # ── 헤더 영역 입력 ──
        # (1,1) 제목: 월 자동 입력
        ws.cell(row=1, column=1).value = f'주간활동서비스 월별 활동계획서({month:02d}월)'

        # (2,4) 작성자, (2,8) 작성일자
        prev_weekday = _last_weekday_prev_month(year, month)
        dow_kr = dow_names_kr[prev_weekday.weekday()]
        date_str = f"{prev_weekday.year}.{prev_weekday.month:02d}.{prev_weekday.day:02d}({dow_kr})"
        ws.cell(row=2, column=4).value = provider
        ws.cell(row=2, column=8).value = date_str

        # (3,4) 수급자(성명), (3,8) 담당 제공인력
        ws.cell(row=3, column=4).value = user_name
        ws.cell(row=3, column=8).value = provider

        # (4,4) 수급시간: ■/□ 표시
        if 수급시간 == 176:
            ws.cell(row=4, column=4).value = '□ 월 132시간 \n■ 월 176시간 '
        else:
            ws.cell(row=4, column=4).value = '■ 월 132시간 \n□ 월 176시간 '

        # (6,4) 총 계획시간
        ws.cell(row=6, column=4).value = f'월 ( {수급시간} )시간'


        # ── 행 수 자동 조정 (시트별 동적 감지) ──
        DATA_START = 9
        # 합계/수식 행을 동적으로 찾기: 행 9부터 스캔하여
        # A열에 "합계"가 있거나 L열에 SUM 수식이 있는 행을 찾는다
        formula_row = None
        for scan_row in range(DATA_START, ws.max_row + 1):
            cell_a = ws.cell(row=scan_row, column=1)
            cell_l = ws.cell(row=scan_row, column=12)
            a_val = cell_a.value
            l_val = cell_l.value if not isinstance(cell_l, MergedCell) else None
            if a_val and '합계' in str(a_val):
                formula_row = scan_row
                break
            if l_val and isinstance(l_val, str) and '=SUM' in l_val.upper():
                formula_row = scan_row
                break
        if formula_row is None:
            formula_row = DATA_START + 21  # fallback
        available_rows = formula_row - DATA_START
        needed_rows = len(working_days)

        # 참조 행(기존 데이터 행)에서 셀 스타일 캡처
        ref_row = DATA_START  # row 9
        ref_styles = {}  # col -> (font, border, alignment)
        for c in range(1, 16):
            cell = ws.cell(row=ref_row, column=c)
            if not isinstance(cell, MergedCell):
                ref_styles[c] = (copy(cell.font), copy(cell.border), copy(cell.alignment))

        # formula_row 이후 행들의 높이 저장
        saved_row_heights = {}
        for r in range(formula_row, ws.max_row + 1):
            h = ws.row_dimensions[r].height
            if h is not None:
                saved_row_heights[r] = h

        if needed_rows > available_rows:
            extra = needed_rows - available_rows
            # formula_row 이후 모든 병합 셀 정보 저장 후 해제
            saved_merges = []
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= formula_row:
                    saved_merges.append((
                        mr.min_row, mr.min_col, mr.max_row, mr.max_col
                    ))
                    ws.unmerge_cells(str(mr))
            ws.insert_rows(formula_row, extra)
            formula_row += extra
            # 저장했던 병합 셀을 이동된 위치로 복원
            for min_r, min_c, max_r, max_c in saved_merges:
                new_min_r = min_r + extra
                new_max_r = max_r + extra
                # 합계 행의 경우 A~K만 병합 (L~O는 수식용)
                if new_min_r == formula_row and max_c >= 12:
                    ws.merge_cells(
                        start_row=new_min_r, start_column=1,
                        end_row=new_max_r, end_column=11
                    )
                else:
                    ws.merge_cells(
                        start_row=new_min_r, start_column=min_c,
                        end_row=new_max_r, end_column=max_c
                    )
            # 이동된 footer 행 높이 복원
            for orig_r, h in saved_row_heights.items():
                ws.row_dimensions[orig_r + extra].height = h
        elif needed_rows < available_rows:
            excess = available_rows - needed_rows
            delete_start = DATA_START + needed_rows
            # 삭제될 데이터 행의 병합 해제 (openpyxl이 자동 정리하지 않음)
            for mr in list(ws.merged_cells.ranges):
                if delete_start <= mr.min_row < formula_row:
                    ws.unmerge_cells(str(mr))
            # footer 병합 저장 후 해제 (delete_rows가 위치를 갱신하지 않으므로)
            saved_merges = []
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row >= formula_row:
                    saved_merges.append((
                        mr.min_row, mr.min_col, mr.max_row, mr.max_col
                    ))
                    ws.unmerge_cells(str(mr))
            ws.delete_rows(delete_start, excess)
            formula_row -= excess
            # footer 병합을 이동된 위치로 복원
            for min_r, min_c, max_r, max_c in saved_merges:
                new_min_r = min_r - excess
                new_max_r = max_r - excess
                if new_min_r == formula_row and max_c >= 12:
                    ws.merge_cells(
                        start_row=new_min_r, start_column=1,
                        end_row=new_max_r, end_column=11
                    )
                else:
                    ws.merge_cells(
                        start_row=new_min_r, start_column=min_c,
                        end_row=new_max_r, end_column=max_c
                    )
            # 이동된 footer 행 높이 복원
            for orig_r, h in saved_row_heights.items():
                ws.row_dimensions[orig_r - excess].height = h

        # ── 데이터 행: 기존 병합 해제 → 스타일 적용 → 재병합 ──
        # 병합 상태에서는 E,F,H,I,K 셀에 테두리를 설정할 수 없으므로
        # 먼저 전부 해제하고, 모든 셀에 스타일 적용 후, 다시 병합한다.
        last_data_row = DATA_START + needed_rows - 1
        for r in range(DATA_START, last_data_row + 1):
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row == r and mr.max_row == r and mr.min_col in (4, 7, 10):
                    ws.unmerge_cells(str(mr))

        # 스타일 적용 (병합 해제 상태이므로 모든 셀에 접근 가능)
        for r in range(DATA_START, last_data_row + 1):
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if c in ref_styles:
                    ref_font, ref_border, ref_align = ref_styles[c]
                    cell.font = copy(ref_font)
                    cell.border = copy(ref_border)
                    cell.alignment = copy(ref_align)

        # 병합 재설정
        for r in range(DATA_START, last_data_row + 1):
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)   # D:F
            ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)   # G:I
            ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)  # J:K

        # ── 데이터 입력 ──
        for i, d in enumerate(working_days):
            row = DATA_START + i

            # A: 날짜, B: 요일, C: 제공자
            for col, val in [(1, d), (2, dow_map.get(d, '')), (3, provider)]:
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = val

            # D: 활동계획
            day_acts = activities.get(d, [])
            lines = []
            if has_오전송영:
                lines.append(오전송영시간)
            lines.extend(day_acts)
            if has_오후송영:
                lines.append(오후송영시간)
            full_text = '\n'.join(lines)

            cell_d = ws.cell(row=row, column=4)
            if not isinstance(cell_d, MergedCell):
                cell_d.value = make_cell_value(full_text)

            ws.row_dimensions[row].height = row_height

        ws.column_dimensions['D'].width = 30

        # ── 수식 업데이트 (행 수 변경 시 범위 보정) ──
        # 합계 행 수식
        for col_letter, col_num in [('L', 12), ('M', 13), ('N', 14), ('O', 15)]:
            ws.cell(row=formula_row, column=col_num).value = (
                f'=SUM({col_letter}{DATA_START}:{col_letter}{last_data_row})'
            )
        # 헤더 수식 (데이터 범위 참조)
        ws.cell(row=4, column=10).value = f'=SUM(L{DATA_START}:L{last_data_row})'
        ws.cell(row=5, column=10).value = f'=SUM(M{DATA_START}:M{last_data_row})'
        ws.cell(row=6, column=10).value = f'=SUM(N{DATA_START}:N{last_data_row})'
        ws.cell(row=7, column=10).value = f'=SUM(O{DATA_START}:O{last_data_row})'
        # 총 계획시간 수식 (합계 행 참조)
        ws.cell(row=5, column=6).value = f'=SUM(L{formula_row}:O{formula_row})'

        results.append({
            'name': user_name,
            'sheet': sheet_name,
            '오전송영': has_오전송영,
            '오후송영': has_오후송영,
            '수급시간': 수급시간,
            'days': len(working_days),
            'formula_row': formula_row,
        })

    # 헤더 영역 병합 셀 테두리 보정
    # openpyxl이 병합 셀(MergedCell)의 스타일을 저장 시 잃어버리는 버그 보정
    # 원본 G4는 right=medium이지만, load→save만 해도 right=thin으로 바뀜
    # RealCell로 교체하여 원본 테두리 복원
    from openpyxl.cell.cell import Cell as RealCell
    for r in results:
        ws = wb[r['sheet']]
        f4 = ws.cell(row=4, column=6)  # F4 (anchor cell, 수정하지 않음)
        # G4: RealCell로 교체하여 원본 스타일 복원
        real_g4 = RealCell(ws, row=4, column=7)
        real_g4.font = copy(f4.font)
        real_g4.fill = copy(f4.fill)
        real_g4.alignment = copy(f4.alignment)
        real_g4.protection = copy(f4.protection)
        real_g4.number_format = f4.number_format
        real_g4.border = Border(
            left=Side(style='thin'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='thin'),
        )
        ws._cells[(4, 7)] = real_g4

    # 저장
    buf = io.BytesIO()
    wb.save(buf)
    output_bytes = buf.getvalue()

    # 수식 보존 검증 (동적 formula_row 사용)
    wb_check = load_workbook(io.BytesIO(output_bytes))
    formulas_ok = True
    for r in results:
        ws = wb_check[r['sheet']]
        fr = r['formula_row']
        for col in [12, 13, 14, 15]:
            val = ws.cell(row=fr, column=col).value
            if val and isinstance(val, str) and val.startswith('='):
                continue
            else:
                formulas_ok = False
    wb_check.close()

    return output_bytes, results, working_days, formulas_ok
